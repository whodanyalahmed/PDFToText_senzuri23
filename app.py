# importing required modules
import PyPDF2
import shutil
import os
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import sys
import os
cur_path = sys.path[0]


def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(__file__)
    return os.path.join(base_path, relative_path)


driverPath = resource_path('chromedriver.exe')
driver = webdriver.Chrome(driverPath)

driver.maximize_window()

driver.implicitly_wait(6)

# Login


def login():
    try:
        driver.find_element_by_id(
            "ctl00_ctl00__dh_UserName").send_keys("Dlaing")
        time.sleep(1)

        driver.find_element_by_id(
            "ctl00_ctl00__dh_Password").send_keys("Hirosushi1")
        time.sleep(1)
        driver.find_element_by_xpath(
            "//html/body/form/div[4]/div[1]/span/div[1]/div[2]/div/span/fieldset/span[3]/input").click()
        time.sleep(1)
    except Exception as e:
        print("error: something went wrong while logging in..." + str(e))

        time.sleep(2)


def go_to_page(url):
    try:
        driver.get(url)
        time.sleep(2)
    except Exception as e:
        print("error: something went wrong while going to page..." + str(e))


def InputKey(key):
    try:
        inputTxt = driver.find_element_by_id(
            "ctl00_ctl00_c_c__txtInvoiceRef__t")
        # clear inputTxt
        inputTxt.clear()
        inputTxt.send_keys(key)
        inputTxt.send_keys(Keys.ENTER)

    except Exception as e:
        print("error: Cant find the element for input " + str(e))


def DownloadPDF(url):
    try:
        driver.find_element_by_id(
            "ctl00_ctl00_c_c__repeaterCustomers_ctl00__btnDownload").click()
    except Exception as e:
        print("error: cant downlaod the file something went wrong..." + str(e))
    finally:
        driver.get(url)


def NamefromPDF(filename, path):

    # creating a pdf file object
    pdfFileObj = open(path+filename, 'rb')

    # creating a pdf reader object
    pdfReader = PyPDF2.PdfFileReader(pdfFileObj)

    # printing number of pages in pdf file
    # print(pdfReader.numPages)

    # creating a page object
    pageObj = pdfReader.getPage(0)

    txt = pageObj.extractText()
    list_txt = txt.split('\n')
    # split txt with new line
    if filename == "CreditNote.pdf":
        print("info : its CreditNote.pdf....")
        main_text = list_txt[30]
    else:
        main_text = list_txt[-7]

    if main_text == "Account No. ":
        main_text = list_txt[61]
        if main_text == "Modules:":
            main_text = list_txt[60]
    elif main_text == "AY":
        main_text = list_txt[-6]
    elif main_text == "Modules:":
        main_text = list_txt[-8]
    else:

        try:

            main_text = list_txt[list_txt.index("Modules:")-1]
        except Exception as e:
            print("info : Modules string is not in the list.. " + str(e))
            main_text = list_txt[-7]

    # closing the pdf file object
    pdfFileObj.close()
    return main_text


def moveFile(filename, path):
    if not os.path.exists(path+'done'):
        os.makedirs(path+'done')
        print('info : created done folder')
    # check if file exist in path
    counter = 1
    filename_new = filename.split('.')
    # check if file exist until it is not exist
    while True:
        counter += 1
        newName = filename_new[0]+'_'+str(counter)+'.'+filename_new[1]
        if os.path.exists(path+"done/"+newName):
            continue
        else:
            break
    # move file to done folder
    shutil.move(path+filename, path+'done/' +
                filename_new[0]+'_'+str(counter)+'.'+filename_new[1])

    print('success : successfully moved ' + newName + ' file to done folder')


if __name__ == '__main__':
    # File name here
    excelFile = "data.xlsx"

    source_file = load_workbook(excelFile, data_only=True)
    # Sheet name here
    sheet = source_file["Sheet2"]

    filename = "invoice.pdf"

    path = "C://Users//dlaing//Downloads//"
    websiteURL = "https://tas.driverhire.co.uk/tas/invoicing/"
    driver.get(websiteURL)
    login()
    go_to_page(websiteURL)
    filename2 = "CreditNote.pdf"
    if os.path.exists(path+filename) or os.path.exists(path+filename2):
        print("info : invoice.pdf already exists moving to folder...")
        moveFile(filename, path)
        moveFile(filename2, path)

    # loop on sheet
    for index, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=3, max_col=3)):
        for cell in row:
            filename = "invoice.pdf"

            num = 'C'+str(index+2)

            if (sheet['D'+str(index+2)].value == None or sheet['D'+str(index+2)].value == 0):
                print(index, sheet['D'+str(index+2)].value, sheet[num].value)

                try:
                    InputKey(sheet[num].value)
                    driver.implicitly_wait(5)
                    DownloadPDF(websiteURL)
                    time.sleep(3)
                    print(filename)
                    if os.path.exists(path+filename):
                        print("info : invoice.pdf is in the path")

                    else:
                        filename = "CreditNote.pdf"
                        print("info : invoice.pdf is not there trying CreditNote.pdf")

                    try:
                        txt = NamefromPDF(filename, path)

                    except Exception as e:
                        print("info : can't find text or pdf...")
                        continue

                    print(txt)
                    sheet['D'+str(index+2)].value = txt
                    source_file.save(excelFile
                                     )
                    try:
                        moveFile(filename, path)
                    except Exception as e:
                        print("error: cant move to folder.. or " + str(e))

                except Exception as e:
                    print("error: something went wrong while inputing key..." + str(e))
                    continue
    # iterate on pair keys and values with index
print("success: Completed...")
